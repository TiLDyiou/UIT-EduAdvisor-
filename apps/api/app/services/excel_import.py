from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

EXAM_REQUIRED_HEADERS = [
    "Mã MH",
    "Tên MH",
    "Mã lớp",
    "Ngày thi",
    "Ca Thi",
    "Phòng Thi",
    "Học kỳ",
    "Năm học",
]

OFFERING_REQUIRED_HEADERS = [
    "term_code",
    "course_code",
    "course_name",
    "credits",
    "section_code",
    "day_of_week",
    "start_period",
    "end_period",
    "room",
]


@dataclass
class PreviewResult:
    ok_rows: list[dict[str, Any]]
    errors: list[dict[str, Any]]
    header: list[str]


def _normalize_header(v: Any) -> str:
    return str(v or "").strip()


def _load_rows(path: str) -> list[tuple[Any, ...]]:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        sheet = wb.active
        return list(sheet.iter_rows(values_only=True))
    finally:
        wb.close()


def validate_xlsx_readable(path: str) -> None:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        _ = wb.active.max_row
    finally:
        wb.close()


def preview_exam_schedule_xlsx(path: str) -> PreviewResult:
    rows = _load_rows(path)
    header_idx = -1
    header: list[str] = []
    for i, row in enumerate(rows[:30]):
        h = [_normalize_header(x) for x in row]
        if "Mã MH" in h and "Ngày thi" in h and "Ca Thi" in h:
            header_idx = i
            header = h
            break
    if header_idx < 0:
        return PreviewResult(
            ok_rows=[], errors=[{"row": 0, "error": "header_not_found"}], header=[]
        )

    missing = [k for k in EXAM_REQUIRED_HEADERS if k not in header]
    if missing:
        return PreviewResult(
            ok_rows=[],
            errors=[{"row": header_idx + 1, "error": "missing_headers", "missing": missing}],
            header=header,
        )

    idx = {k: header.index(k) for k in EXAM_REQUIRED_HEADERS}
    ok_rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for i, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        vals = [row[j] if j < len(row) else None for j in range(len(header))]
        if not any(v is not None and str(v).strip() for v in vals):
            continue
        raw = {k: vals[idx[k]] for k in EXAM_REQUIRED_HEADERS}
        try:
            course_code = str(raw["Mã MH"]).strip().upper()
            course_name = str(raw["Tên MH"]).strip()
            section_code = str(raw["Mã lớp"]).strip()
            room = str(raw["Phòng Thi"]).strip()
            if not course_code or not course_name or not section_code:
                raise ValueError("empty_required_cells")

            dt = _parse_exam_date(raw["Ngày thi"])
            ca = int(str(raw["Ca Thi"]).strip())
            if ca < 1:
                raise ValueError("invalid_ca_thi")
            start_t, end_t = _ca_to_time(ca)
            term_code = f"{str(raw['Năm học']).strip()}-{str(raw['Học kỳ']).strip()}"
            ok_rows.append(
                {
                    "course_code": course_code,
                    "course_name": course_name,
                    "section_code": section_code,
                    "exam_date": dt.date().isoformat(),
                    "start_time": start_t.strftime("%H:%M:%S"),
                    "end_time": end_t.strftime("%H:%M:%S"),
                    "room": room or None,
                    "term_code": term_code,
                }
            )
        except Exception as exc:
            errors.append({"row": i, "error": str(exc)})
    return PreviewResult(ok_rows=ok_rows, errors=errors, header=header)


def _parse_exam_date(v: Any) -> datetime:
    if isinstance(v, datetime):
        return v
    txt = str(v or "").strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(txt, fmt)
        except ValueError:
            pass
    raise ValueError("invalid_exam_date")


def _ca_to_time(ca: int) -> tuple[time, time]:
    # Mapping tối thiểu để chạy production local; có thể đổi theo quy định chính thức.
    mapping = {
        1: (time(7, 30), time(9, 30)),
        2: (time(9, 45), time(11, 45)),
        3: (time(13, 30), time(15, 30)),
        4: (time(15, 45), time(17, 45)),
    }
    if ca not in mapping:
        raise ValueError("unsupported_ca_thi")
    return mapping[ca]


def preview_course_offerings_file(path: str) -> PreviewResult:
    if Path(path).suffix.lower() != ".xlsx":
        return PreviewResult(
            ok_rows=[], errors=[{"row": 0, "error": "only_xlsx_supported_now"}], header=[]
        )
    rows = _load_rows(path)
    if not rows:
        return PreviewResult(ok_rows=[], errors=[{"row": 0, "error": "empty_file"}], header=[])
    header = [_normalize_header(x) for x in rows[0]]
    missing = [k for k in OFFERING_REQUIRED_HEADERS if k not in header]
    if missing:
        return PreviewResult(
            ok_rows=[],
            errors=[{"row": 1, "error": "missing_headers", "missing": missing}],
            header=header,
        )
    idx = {k: header.index(k) for k in OFFERING_REQUIRED_HEADERS}
    ok_rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for i, row in enumerate(rows[1:], start=2):
        vals = [row[j] if j < len(row) else None for j in range(len(header))]
        if not any(v is not None and str(v).strip() for v in vals):
            continue
        try:
            course_code = str(vals[idx["course_code"]]).strip().upper()
            term_code = str(vals[idx["term_code"]]).strip()
            course_name = str(vals[idx["course_name"]]).strip()
            credits = int(vals[idx["credits"]])
            if not course_code or not term_code or not course_name or credits <= 0:
                raise ValueError("invalid_required_fields")
            ok_rows.append(
                {
                    "term_code": term_code,
                    "course_code": course_code,
                    "course_name": course_name,
                    "credits": credits,
                    "section_code": str(vals[idx["section_code"]] or "").strip(),
                    "day_of_week": int(vals[idx["day_of_week"]])
                    if vals[idx["day_of_week"]] is not None
                    else None,
                    "start_period": int(vals[idx["start_period"]])
                    if vals[idx["start_period"]] is not None
                    else None,
                    "end_period": int(vals[idx["end_period"]])
                    if vals[idx["end_period"]] is not None
                    else None,
                    "room": str(vals[idx["room"]]).strip() or None,
                }
            )
        except Exception as exc:
            errors.append({"row": i, "error": str(exc)})
    return PreviewResult(ok_rows=ok_rows, errors=errors, header=header)
